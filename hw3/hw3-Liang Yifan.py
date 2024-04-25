import os
import sys
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook


#homework3 by Liang Yifan
gravity=9.8
g=gravity

class Movement:
    def __init__(self,vx0,vy0,time=0,timestep=1):
        self.x0=0
        self.y0=0
        self.t=np.arange(0,time+timestep,timestep)
        self.vx=vx0-self.t*g
        self.vy=vy0+0*self.t
        self.x=self.x0+vx0*self.t-0.5*g*self.t*self.t
        self.y=self.y0+vy0*self.t

    def print_result(self):
        print('t  x  y  vx  vy')
        for i in range(self.t.size):
            print(self.t[i],self.x[i],self.y[i],self.vx[i],self.vy[i])

    def save_data(self):
        wb=Workbook()
        #wb.save('hw3_data.xlsx')
        sheet = wb['Sheet']
        sheet.cell(row=1, column=1).value = 't'
        sheet.cell(row=1, column=2).value = 'x'
        sheet.cell(row=1, column=3).value = 'y'
        sheet.cell(row=1, column=4).value = 'vx'
        sheet.cell(row=1, column=5).value = 'vy'
        for i in range(self.t.size):
            sheet.cell(row=i+2, column=1).value=self.t[i]
            sheet.cell(row=i+2, column=2).value=self.x[i]
            sheet.cell(row=i+2, column=3).value=self.y[i]
            sheet.cell(row=i+2, column=4).value=self.vx[i]
            sheet.cell(row=i+2, column=5).value=self.vy[i]
        wb.save('hw3_data.xlsx')

    def plot_data(self):
        plt.plot(self.y,self.x)
        plt.xlabel('y')
        plt.ylabel('x')
        plt.savefig('hw3.png')
        plt.show()


if __name__=='__main__':
    move=Movement(50,50,10)
    move.print_result()
    move.save_data()
    move.plot_data()
